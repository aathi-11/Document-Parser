import hashlib
import json
import logging
from pathlib import Path

import openpyxl
import pandas as pd

logger = logging.getLogger(__name__)


def _hash_file(path: Path) -> str:
    hasher = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            hasher.update(chunk)
    return hasher.hexdigest()


def _build_upload_manifest(uploads: list[Path]) -> dict[str, str]:
    return {upload.name: _hash_file(upload) for upload in uploads}


def _load_manifest(path: Path) -> dict[str, str]:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}

def preprocess_tabular_files(session_dir: Path) -> Path:
    """
    Find all Excel and CSV files in session_dir/uploads,
    preprocess and clean their sheets, and save them as standardized CSV files
    in session_dir/cleaned_csvs.
    """
    upload_dir = session_dir / "uploads"
    cleaned_dir = session_dir / "cleaned_csvs"
    
    if not upload_dir.exists():
        return cleaned_dir
        
    cleaned_dir.mkdir(parents=True, exist_ok=True)
    
    manifest_path = cleaned_dir / "manifest.json"
    uploads = [p for p in upload_dir.iterdir() if p.suffix.lower() in {".xlsx", ".xlsm", ".xls", ".csv"}]
    cleaned_files = [p for p in cleaned_dir.iterdir() if p.suffix.lower() == ".csv"]
    current_manifest: dict[str, str] = {}
    if uploads:
        try:
            current_manifest = _build_upload_manifest(uploads)
        except Exception as exc:
            logger.error(f"Failed to build upload manifest: {exc}")
            current_manifest = {}

    if uploads and cleaned_files and current_manifest and manifest_path.exists():
        cached_manifest = _load_manifest(manifest_path)
        if cached_manifest == current_manifest:
            logger.info("Cleaned CSVs are up to date, skipping preprocessing.")
            return cleaned_dir
            
    # Process all files
    had_errors = False
    for filepath in upload_dir.iterdir():
        ext = filepath.suffix.lower()
        if ext not in {".xlsx", ".xlsm", ".xls", ".csv"}:
            continue
            
        try:
            if ext == ".csv":
                # Copy CSV or clean it slightly
                df = pd.read_csv(filepath)
                df.columns = df.columns.str.strip()
                df = df.apply(lambda x: x.str.strip() if hasattr(x, 'str') else x)
                # Drop completely blank rows
                df = df[df.apply(lambda row: any(str(c).strip() for c in row), axis=1)]
                # Drop summary/total rows if they exist (checking first 2 columns)
                if len(df.columns) > 0:
                    cols_to_check = df.columns[:2]
                    total_mask = pd.Series(False, index=df.index)
                    for col in cols_to_check:
                        val_lower = df[col].astype(str).str.lower().str.strip()
                        mask = val_lower.isin([
                            'total', 'totals', 'portfolio total', 'total portfolio', 'global total', 'grand total', 'aggregate'
                        ]) | val_lower.str.startswith('total ') | val_lower.str.startswith('totals ') | (val_lower == 'global total')
                        total_mask = total_mask | mask
                    df = df[~total_mask]
                dest_path = cleaned_dir / filepath.name
                df.to_csv(dest_path, index=False)
                logger.info(f"Preprocessed CSV: {filepath.name}")
            else:
                # Excel file
                _preprocess_excel(filepath, cleaned_dir)
        except Exception as e:
            had_errors = True
            logger.error(f"Error preprocessing {filepath.name}: {e}", exc_info=True)

    if uploads and current_manifest and not had_errors:
        try:
            manifest_path.write_text(json.dumps(current_manifest, indent=2), encoding="utf-8")
        except OSError as exc:
            logger.error(f"Failed to write manifest: {exc}")

    return cleaned_dir

def _preprocess_excel(filepath: Path, dest_dir: Path) -> None:
    ext = filepath.suffix.lower()
    
    if ext == ".xls":
        xls = pd.ExcelFile(str(filepath), engine="xlrd")
        sheet_names = xls.sheet_names
        is_legacy = True
    else:
        wb = openpyxl.load_workbook(str(filepath), data_only=True)
        sheet_names = wb.sheetnames
        is_legacy = False
        
    for name in sheet_names:
        if name.upper() == 'README' or name.upper() == 'DATA DICTIONARY':
            continue
            
        raw_rows = []
        if is_legacy:
            df_sheet = xls.parse(name, header=None, dtype=str).fillna("")
            raw_rows = df_sheet.values.tolist()
        else:
            ws = wb[name]
            for r in ws.iter_rows(values_only=True):
                raw_rows.append([str(c) if c is not None else "" for c in r])
                
        if not raw_rows:
            continue
            
        # Trim blank trailing columns
        max_col = max(
            (max((i + 1 for i, c in enumerate(row) if c.strip()), default=0) for row in raw_rows),
            default=0,
        )
        raw_rows = [row[:max_col] for row in raw_rows]
        
        # Header row detection heuristic: scan first 10 rows
        row_counts = []
        for idx, row in enumerate(raw_rows[:10]):
            non_empty = sum(1 for c in row if c.strip())
            row_counts.append((idx, non_empty))
            
        max_non_empty = max(count for idx, count in row_counts) if row_counts else 0
        
        header_idx = 0
        for idx, count in row_counts:
            # We want a row that contains a good portion of cells filled
            if count >= 3 and count >= max_non_empty * 0.5:
                header_idx = idx
                break
                
        headers = [h.strip() for h in raw_rows[header_idx]]
        data_rows = raw_rows[header_idx + 1:]
        
        # Build DataFrame
        df = pd.DataFrame(data_rows, columns=headers)
        df = df.apply(lambda x: x.str.strip() if hasattr(x, 'str') else x)
        
        # Drop completely blank rows
        df = df[df.apply(lambda row: any(str(c).strip() for c in row), axis=1)]
        
        # Drop summary/total rows if they exist (checking first 2 columns)
        if len(df.columns) > 0:
            cols_to_check = df.columns[:2]
            total_mask = pd.Series(False, index=df.index)
            for col in cols_to_check:
                val_lower = df[col].astype(str).str.lower().str.strip()
                mask = val_lower.isin([
                    'total', 'totals', 'portfolio total', 'total portfolio', 'global total', 'grand total', 'aggregate'
                ]) | val_lower.str.startswith('total ') | val_lower.str.startswith('totals ') | (val_lower == 'global total')
                total_mask = total_mask | mask
            df = df[~total_mask]
            
        # Specialized Actuarial Loss Triangle cleanups
        if 'loss_triangle' in name.lower() or 'loss triangle' in name.lower():
            # Look at the first column (typically Accident Year)
            first_col = df.columns[0]
            # Drop the bottom rows (development factors/averages) that aren't numeric years
            df['__temp_year'] = pd.to_numeric(df[first_col], errors='coerce')
            # Keep only rows where Accident Year is a valid number and Earned Premium is present
            earned_premium_col = df.columns[1] if len(df.columns) > 1 else None
            if earned_premium_col:
                df = df[df['__temp_year'].notna() & (df[earned_premium_col].str.strip() != "")]
            else:
                df = df[df['__temp_year'].notna()]
            df = df.drop(columns=['__temp_year'])
            
            # Convert numeric columns to float/int directly
            for col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')
                
        # Save sheet as CSV
        csv_path = dest_dir / f"{name}.csv"
        df.to_csv(csv_path, index=False)
        logger.info(f"Preprocessed sheet '{name}' to {csv_path.name}")
