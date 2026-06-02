import logging
import re
from pathlib import Path
from typing import Any, Dict

import pandas as pd
import openpyxl
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

from app.services.tabular_preprocessor import preprocess_tabular_files
from app.services.tabular_query import sanitize_df_name

logger = logging.getLogger(__name__)


def _select_target_df(
    question: str,
    preloaded_dfs: Dict[str, pd.DataFrame],
    csv_files: list[Path],
) -> str:
    lowered = question.lower()
    for csv_file in csv_files:
        stem = csv_file.stem.lower()
        if stem and stem in lowered:
            var_name = sanitize_df_name(csv_file.name)
            if var_name in preloaded_dfs:
                return var_name
    for var_name in preloaded_dfs.keys():
        if var_name.lower() in lowered:
            return var_name
    return next(iter(preloaded_dfs.keys()))


def _select_column_by_question(question: str, columns: list[str]) -> str | None:
    lowered = question.lower()
    for col in columns:
        col_name = str(col).lower()
        if col_name and col_name in lowered:
            return col
    return None


def _coerce_numeric(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce")


def _coerce_datetime(series: pd.Series) -> pd.Series:
    return pd.to_datetime(series, errors="coerce", utc=False)


def _get_numeric_columns(df: pd.DataFrame) -> list[str]:
    numeric_cols = list(df.select_dtypes(include="number").columns)
    for col in df.columns:
        if col in numeric_cols:
            continue
        values = _coerce_numeric(df[col])
        if values.notna().mean() >= 0.6:
            numeric_cols.append(col)
    return numeric_cols


def _get_date_column(df: pd.DataFrame) -> str | None:
    for col in df.columns:
        col_name = str(col).lower()
        if "date" in col_name or "time" in col_name:
            parsed = _coerce_datetime(df[col])
            if parsed.notna().mean() >= 0.6:
                return col
    for col in df.columns:
        parsed = _coerce_datetime(df[col])
        if parsed.notna().mean() >= 0.6:
            return col
    return None


def _get_category_column(
    df: pd.DataFrame,
    numeric_cols: list[str],
    date_col: str | None,
) -> str | None:
    candidate_cols = [
        col for col in df.columns if col not in numeric_cols and col != date_col
    ]
    object_cols = [col for col in candidate_cols if df[col].dtype == "object"]
    candidates = object_cols or candidate_cols
    best = None
    best_unique = None
    for col in candidates:
        unique_count = df[col].nunique(dropna=True)
        if unique_count < 2:
            continue
        if unique_count <= 15:
            if best_unique is None or unique_count < best_unique:
                best = col
                best_unique = unique_count
    if best is not None:
        return best
    return candidates[0] if candidates else None


def _write_table(ws, df: pd.DataFrame, start_row: int, start_col: int) -> tuple[int, int, int, int]:
    for r_idx, row in enumerate(
        dataframe_to_rows(df, index=False, header=True), start=start_row
    ):
        for c_idx, value in enumerate(row, start=start_col):
            ws.cell(row=r_idx, column=c_idx, value=value)
    end_row = start_row + len(df)
    end_col = start_col + len(df.columns) - 1
    return start_row, start_col, end_row, end_col


def _build_dashboard_workbook(df: pd.DataFrame, question: str) -> openpyxl.Workbook | None:
    if df.empty:
        return None

    numeric_cols = _get_numeric_columns(df)
    date_col = _get_date_column(df)
    category_col = _get_category_column(df, numeric_cols, date_col)

    if numeric_cols:
        preferred_numeric = _select_column_by_question(question, numeric_cols)
        numeric_col = preferred_numeric or numeric_cols[0]
    else:
        numeric_col = None

    if category_col:
        preferred_category = _select_column_by_question(question, [category_col])
        category_col = preferred_category or category_col

    max_categories = 10
    count_df = pd.DataFrame()
    sum_df = pd.DataFrame()
    time_df = pd.DataFrame()

    if category_col:
        category_series = df[category_col].astype(str).str.strip()
        count_df = (
            category_series.value_counts()
            .head(max_categories)
            .reset_index()
        )
        count_df.columns = [category_col, "Count"]

        if numeric_col:
            value_series = _coerce_numeric(df[numeric_col])
            sum_df = (
                pd.DataFrame({category_col: category_series, numeric_col: value_series})
                .dropna(subset=[numeric_col])
                .groupby(category_col, dropna=True)[numeric_col]
                .sum()
                .sort_values(ascending=False)
                .head(max_categories)
                .reset_index()
            )
            sum_df.columns = [category_col, f"Total {numeric_col}"]

    if date_col:
        date_series = _coerce_datetime(df[date_col])
        if numeric_col:
            value_series = _coerce_numeric(df[numeric_col])
            time_df = pd.DataFrame({"Period": date_series, "Value": value_series})
            time_df = time_df.dropna(subset=["Period", "Value"])
        else:
            time_df = pd.DataFrame({"Period": date_series})
            time_df = time_df.dropna(subset=["Period"])
            time_df["Value"] = 1

        if not time_df.empty:
            time_df["Period"] = time_df["Period"].dt.to_period("M").dt.to_timestamp()
            time_df = (
                time_df.groupby("Period", as_index=False)["Value"]
                .sum()
                .sort_values("Period")
            )
            time_df["Period"] = time_df["Period"].dt.strftime("%Y-%m")
            time_df = time_df.head(12)

    has_any = not count_df.empty or not sum_df.empty or not time_df.empty
    if not has_any:
        return None

    result_wb = openpyxl.Workbook()
    result_wb.remove(result_wb.active)

    dash_ws = result_wb.create_sheet("Dashboard")
    data_ws = result_wb.create_sheet("ChartData")
    data_ws.sheet_state = "hidden"

    tables: dict[str, tuple[int, int, int, int]] = {}
    row_cursor = 1

    if not count_df.empty:
        tables["count"] = _write_table(data_ws, count_df, row_cursor, 1)
        row_cursor = tables["count"][2] + 3

    if not sum_df.empty:
        tables["sum"] = _write_table(data_ws, sum_df, row_cursor, 1)
        row_cursor = tables["sum"][2] + 3

    if not time_df.empty:
        tables["time"] = _write_table(data_ws, time_df, row_cursor, 1)

    chart_positions = ["A1", "K1", "A20"]
    chart_index = 0

    if "count" in tables:
        start_row, start_col, end_row, end_col = tables["count"]
        data_ref = Reference(
            data_ws,
            min_col=start_col + 1,
            max_col=end_col,
            min_row=start_row,
            max_row=end_row,
        )
        cat_ref = Reference(
            data_ws,
            min_col=start_col,
            min_row=start_row + 1,
            max_row=end_row,
        )
        chart = BarChart()
        chart.title = f"{category_col} Count" if category_col else "Count by Category"
        chart.style = 10
        chart.width = 20
        chart.height = 12
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cat_ref)
        dash_ws.add_chart(chart, chart_positions[chart_index])
        chart_index += 1

    if "sum" in tables:
        start_row, start_col, end_row, _ = tables["sum"]
        data_ref = Reference(
            data_ws,
            min_col=start_col + 1,
            max_col=start_col + 1,
            min_row=start_row,
            max_row=end_row,
        )
        cat_ref = Reference(
            data_ws,
            min_col=start_col,
            min_row=start_row + 1,
            max_row=end_row,
        )
        chart = PieChart()
        chart.title = f"{numeric_col} Distribution" if numeric_col else "Distribution"
        chart.style = 10
        chart.width = 20
        chart.height = 12
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cat_ref)
        dash_ws.add_chart(chart, chart_positions[chart_index])
        chart_index += 1

    if "time" in tables:
        start_row, start_col, end_row, end_col = tables["time"]
        data_ref = Reference(
            data_ws,
            min_col=start_col + 1,
            max_col=end_col,
            min_row=start_row,
            max_row=end_row,
        )
        cat_ref = Reference(
            data_ws,
            min_col=start_col,
            min_row=start_row + 1,
            max_row=end_row,
        )
        chart = LineChart()
        chart.title = "Trend Over Time"
        chart.style = 10
        chart.width = 20
        chart.height = 12
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cat_ref)
        chart.smooth = True
        dash_ws.add_chart(chart, chart_positions[chart_index])

    dash_ws.sheet_view.showGridLines = False
    return result_wb


def is_visualization_question(question: str) -> bool:
    phrases = [
        "dashboard", "visuali", "chart", "graph", "plot",
        "bar chart", "line chart", "pie chart", "area chart",
        "column chart", "scatter chart", "chart in excel",
        "charts in excel", "excel chart", "excel dashboard",
        "add chart", "create chart", "generate chart",
        "insert chart", "draw chart", "make chart",
        "add graph", "create graph", "visualise", "visualize"
    ]
    lowered = question.lower()
    return any(phrase in lowered for phrase in phrases)


def run_excel_visualization(session_dir: Path, question: str, base_url: str, model: str) -> Dict[str, Any]:
    # Part A — Load DataFrames
    cleaned_dir = preprocess_tabular_files(session_dir)
    if not cleaned_dir.exists():
        return {"handled": False}

    csv_files = [p for p in cleaned_dir.iterdir() if p.suffix.lower() == ".csv"]
    if not csv_files:
        return {"handled": False}

    preloaded_dfs = {}

    for csv_file in sorted(csv_files):
        try:
            df = pd.read_csv(csv_file)
            var_name = sanitize_df_name(csv_file.name)
            preloaded_dfs[var_name] = df

        except Exception as e:
            logger.error(f"Error preloading/parsing {csv_file.name}: {e}")

    # Part B — Build dashboard deterministically to avoid invalid chart ranges
    success = False
    result_wb = None
    try:
        target_df_name = _select_target_df(question, preloaded_dfs, csv_files)
        target_df = preloaded_dfs[target_df_name]
        result_wb = _build_dashboard_workbook(target_df, question)
        success = result_wb is not None
    except Exception as exc:
        logger.error(f"Error while building visualization dashboard: {exc}")

    # Part C — Save the workbook
    if success and result_wb is not None:
        # Always inject Original Data as the first sheet
        original_inserted = False
        uploads_dir = session_dir / "uploads"
        if uploads_dir.exists():
            excel_files = sorted([
                f for f in uploads_dir.iterdir()
                if f.suffix.lower() in {".xlsx", ".xls", ".csv"}
            ])
            if excel_files:
                try:
                    src = excel_files[0]
                    if src.suffix.lower() == ".csv":
                        orig_df = pd.read_csv(src)
                    else:
                        orig_df = pd.read_excel(src, engine="openpyxl")

                    orig_ws = result_wb.create_sheet("Original Data", 0)
                    orig_ws.sheet_properties.tabColor = "1F3864"
                    for row in dataframe_to_rows(orig_df, index=False, header=True):
                        orig_ws.append(row)
                    for cell in orig_ws[1]:
                        cell.font = Font(bold=True, color="FFFFFF", size=11)
                        cell.fill = PatternFill("solid", fgColor="1F3864")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    orig_ws.freeze_panes = "A2"
                    orig_ws.auto_filter.ref = orig_ws.dimensions
                    original_inserted = True
                except Exception as e:
                    logger.warning(f"Could not insert original data sheet: {e}")

        outputs_dir = session_dir / "edited_outputs"
        outputs_dir.mkdir(parents=True, exist_ok=True)

        # Generate filename: take first 6 words of question, join with underscores, strip non-alphanumeric, append _dashboard.xlsx
        words = question.split()[:6]
        joined = "_".join(words)
        sanitized = re.sub(r"[^A-Za-z0-9_]", "", joined)
        filename = f"{sanitized.lower()}_dashboard.xlsx"
        output_path = outputs_dir / filename

        result_wb.save(str(output_path))
        sheet_names = result_wb.sheetnames
        total_charts = sum(len(ws._charts) for ws in result_wb.worksheets)

        return {
            "handled": True,
            "filename": output_path.name,
            "sheet_names": sheet_names,
            "total_charts": total_charts,
            "has_dashboard": "Dashboard" in sheet_names,
            "has_original": original_inserted,
        }

    return {"handled": False}


def get_visualization_file_path(session_dir: Path, filename: str) -> Path | None:
    sanitised_filename = re.sub(r"[^A-Za-z0-9_\-.]", "", Path(filename).name)
    target_path = session_dir / "edited_outputs" / sanitised_filename
    if target_path.exists():
        return target_path
    return None
