import contextlib
import io
import logging
import re
import concurrent.futures
from pathlib import Path
from typing import Any, Dict

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from app.services.tabular_preprocessor import preprocess_tabular_files
from app.services.tabular_query import (
    _ensure_safe_code,
    _run_with_timeout,
    sanitize_df_name,
)

logger = logging.getLogger(__name__)

_EDIT_TRIGGERS = [
    "add row",
    "add a row",
    "append row",
    "append a row",
    "insert row",
    "insert a row",
    "new row",
    "blank row",
    "empty row",
    "emptyrow",
    "create row",
    "add rows",
    "insert rows",
    "pivot",
    "pivot table",
    "add column",
    "add a column",
    "new column",
    "create column",
    "calculate column",
    "append column",
    "insert column",
    "rename column",
    "drop column",
    "remove column",
    "delete column",
    "fill column",
    "update column",
    "modify",
    "edit the",
    "transform",
    "reshape",
    "melt",
    "unpivot",
    "transpose",
    "group by and pivot",
    "crosstab",
]

_BLANK_ROW_RE = re.compile(
    r"\b(?:add|append|insert|create)\s+(?:a\s*)?(?:blank|empty)\s*row\b",
    re.IGNORECASE,
)
_NAMED_VALUE_RE = re.compile(r"\bnamed\s+([^.,;\n]+)", re.IGNORECASE)


def is_edit_question(question: str) -> bool:
    lowered = question.lower()
    return any(trigger in lowered for trigger in _EDIT_TRIGGERS)


def is_combined_download_question(question: str) -> bool:
    lowered = question.lower()
    pivot_phrases = [
        "pivot",
        "pivot table",
        "group by",
        "crosstab",
        "summarise",
        "summarize",
        "breakdown",
        "break down",
        "aggregate",
        "aggregation",
    ]
    download_phrases = [
        "download",
        "export",
        "save",
        "file",
        "excel",
        "xlsx",
        "sheet",
        "get file",
        "give me",
        "generate file",
        "create file",
    ]
    return any(phrase in lowered for phrase in pivot_phrases) and any(
        phrase in lowered for phrase in download_phrases
    )


def _is_blank_row_request(question: str) -> bool:
    return bool(_BLANK_ROW_RE.search(question))


def _extract_named_value(question: str) -> str | None:
    match = _NAMED_VALUE_RE.search(question)
    if not match:
        return None
    value = match.group(1).strip().strip("\"'")
    if not value:
        return None
    value = re.split(r"\b(?:in|at|for|with|on|into|to|of)\b", value, maxsplit=1)[0].strip()
    return value or None


def _select_target_df(
    question: str,
    preloaded_dfs: Dict[str, pd.DataFrame],
    csv_files: list[Path],
) -> str:
    lowered = question.lower()
    for csv_file in csv_files:
        stem = csv_file.stem.lower()
        if stem and stem in lowered:
            var_name = sanitize_df_name(csv_file.name)
            if var_name in preloaded_dfs:
                return var_name
    for var_name in preloaded_dfs.keys():
        if var_name.lower() in lowered:
            return var_name
    return next(iter(preloaded_dfs.keys()))


def _find_name_column(df: pd.DataFrame) -> str | None:
    for col in df.columns:
        col_name = str(col).strip().lower()
        if "name" in col_name:
            return col
    for col in df.columns:
        if df[col].dtype == "object":
            return col
    return None


def _append_blank_row(df: pd.DataFrame, name_value: str | None) -> pd.DataFrame:
    new_row = {col: None for col in df.columns}
    if name_value:
        name_col = _find_name_column(df)
        if name_col is not None:
            new_row[name_col] = name_value
    return pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)


def _write_df_to_sheet(ws, df: pd.DataFrame, sheet_title: str) -> None:
    for row in dataframe_to_rows(df, index=True, header=True):
        ws.append(row)

    ws.title = sheet_title

    header_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_alignment = Alignment(horizontal="center", vertical="center")
    header_border = Border(
        left=Side(style="thin", color="FFFFFF"),
        right=Side(style="thin", color="FFFFFF"),
        top=Side(style="thin", color="FFFFFF"),
        bottom=Side(style="thin", color="FFFFFF"),
    )

    for cell in ws[1]:
        if cell.value is None:
            continue
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border

    data_font = Font(name="Calibri", size=10)
    data_alignment = Alignment(horizontal="left", vertical="center")
    data_border = Border(
        left=Side(style="thin", color="D0D7E8"),
        right=Side(style="thin", color="D0D7E8"),
        top=Side(style="thin", color="D0D7E8"),
        bottom=Side(style="thin", color="D0D7E8"),
    )
    even_fill = PatternFill("solid", fgColor="EBF0FA")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row_index = row[0].row
        fill = even_fill if row_index % 2 == 0 else None
        for cell in row:
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = data_border
            if fill:
                cell.fill = fill

    for col_index, column_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column_cells:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_index)].width = min(max_len + 4, 45)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _add_charts_to_sheet(
    wb: openpyxl.Workbook,
    pivot_ws,
    pivot_df: pd.DataFrame,
) -> None:
    dash_ws = wb.create_sheet("Dashboard")
    dash_ws.sheet_properties.tabColor = "2563EB"

    for col_index in range(1, 27):
        dash_ws.column_dimensions[get_column_letter(col_index)].width = 3

    max_col = min(pivot_ws.max_column, 5)
    max_row = min(pivot_ws.max_row, 20)
    if max_col >= 2 and max_row >= 2:
        data_ref = Reference(
            pivot_ws,
            min_col=2,
            max_col=max_col,
            min_row=1,
            max_row=max_row,
        )
        category_ref = Reference(
            pivot_ws,
            min_col=1,
            min_row=2,
            max_row=max_row,
        )

        chart1 = BarChart()
        chart1.type = "col"
        chart1.title = "Summary by Category"
        chart1.style = 10
        chart1.width = 22
        chart1.height = 14
        chart1.add_data(data_ref, titles_from_data=True)
        chart1.set_categories(category_ref)
        chart1.shape = 4
        dash_ws.add_chart(chart1, "B2")

        numeric_cols = list(pivot_df.select_dtypes(include="number").columns)
        if numeric_cols and len(pivot_df) > 3:
            chart2 = LineChart()
            chart2.title = "Trend Overview"
            chart2.style = 10
            chart2.width = 22
            chart2.height = 14
            chart2.add_data(data_ref, titles_from_data=True)
            chart2.set_categories(category_ref)
            chart2.grouping = "standard"
            chart2.smooth = True
            dash_ws.add_chart(chart2, "M2")

        if len(pivot_df) >= 2 and numeric_cols:
            first_numeric = numeric_cols[0]
            numeric_index = list(pivot_df.columns).index(first_numeric) + 1
            pie_data_ref = Reference(
                pivot_ws,
                min_col=numeric_index,
                max_col=numeric_index,
                min_row=1,
                max_row=min(pivot_ws.max_row, 10),
            )
            pie_cat_ref = Reference(
                pivot_ws,
                min_col=1,
                min_row=2,
                max_row=min(pivot_ws.max_row, 10),
            )
            chart3 = PieChart()
            chart3.title = "Distribution"
            chart3.style = 10
            chart3.width = 22
            chart3.height = 14
            chart3.add_data(pie_data_ref, titles_from_data=True)
            chart3.set_categories(pie_cat_ref)
            dash_ws.add_chart(chart3, "B30")

    dash_ws.sheet_view.showGridLines = False


def run_excel_combined(
    session_dir: Path,
    question: str,
    base_url: str,
    model: str,
) -> Dict[str, Any]:
    cleaned_dir = preprocess_tabular_files(session_dir)
    if not cleaned_dir.exists():
        return {"handled": False}

    csv_files = [p for p in cleaned_dir.iterdir() if p.suffix.lower() == ".csv"]
    if not csv_files:
        return {"handled": False}

    preloaded_dfs: Dict[str, pd.DataFrame] = {}
    schema_parts: list[str] = []

    for csv_file in sorted(csv_files):
        try:
            df = pd.read_csv(csv_file)
        except Exception as exc:
            logger.error(f"Error preloading/parsing {csv_file.name}: {exc}")
            continue

        var_name = sanitize_df_name(csv_file.name)
        preloaded_dfs[var_name] = df
        cols = df.columns.tolist()
        sample_rows = df.head(2).to_string(index=False)

        schema_parts.append(
            f"- DataFrame variable: `{var_name}` (sourced from '{csv_file.name}')\n"
            f"  Columns: {cols}\n"
            f"  Sample Data:\n"
            f"  {sample_rows}"
        )

    if not preloaded_dfs:
        return {"handled": False}

    primary_df = max(preloaded_dfs.values(), key=len)
    schema_str = "\n\n".join(schema_parts)

    prompt = f"""You are a professional Python data analyst. The user wants a pivot table output.

The following pandas DataFrames are pre-loaded in the environment:
{schema_str}

Instructions:
- Write Python Pandas code that produces a pivot table stored in a variable called exactly result_df.
- Use pd.pivot_table or df.groupby to compute the pivot.
- The result must be a pandas DataFrame.
- Do NOT use pd.read_csv or read any files from disk.
- Do NOT import anything (pd and np are already available).
- Do NOT call plt.show() or any plotting functions.
- Print result_df.to_string() to stdout at the end.
- Return only the Python code block starting with ```python and ending with ```.

User instruction: {question}
Python Code:"""

    from app.services.ollama_client import ollama_chat

    messages = [{"role": "user", "content": prompt}]
    success = False
    result_df: pd.DataFrame | None = None

    for attempt in range(1, 4):
        code = ""
        try:
            text = ollama_chat(base_url, model, messages)

            code = text
            if "```python" in text:
                code = text.split("```python")[1].split("```")[0]

            _ensure_safe_code(code)

            exec_globals = {"pd": pd, "np": np, **preloaded_dfs}
            exec_locals: Dict[str, Any] = {}

            stdout_buffer = io.StringIO()
            with contextlib.redirect_stdout(stdout_buffer):
                _run_with_timeout(lambda: exec(code, exec_globals, exec_locals), 20)

            result_df = exec_locals.get("result_df")
            if result_df is None:
                result_df = exec_globals.get("result_df")

            if isinstance(result_df, pd.DataFrame):
                success = True
                break

            raise RuntimeError("result_df was not set to a pandas DataFrame.")
        except Exception:
            import traceback

            error_str = traceback.format_exc()
            logger.warning(
                f"Excel combined execution failed on attempt {attempt}. Error:\n{error_str}"
            )
            if code:
                messages.append({"role": "assistant", "content": f"```python\n{code}\n```"})
            error_msg = (
                "The code execution failed with the following traceback/error:\n"
                f"{error_str.strip()}\n\n"
                "Please correct the code. Ensure you use the correct preloaded DataFrame variables "
                "and write correct Pandas code. Return ONLY the executable python block."
            )
            messages.append({"role": "user", "content": error_msg})

    if not success or result_df is None:
        return {"handled": False}

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_original = wb.create_sheet("Original Data")
    ws_original.sheet_properties.tabColor = "1F3864"
    _write_df_to_sheet(ws_original, primary_df, "Original Data")

    ws_pivot = wb.create_sheet("Pivot Table")
    ws_pivot.sheet_properties.tabColor = "2E7D32"
    pivot_df_reset = result_df.reset_index()
    _write_df_to_sheet(ws_pivot, pivot_df_reset, "Pivot Table")

    _add_charts_to_sheet(wb, ws_pivot, pivot_df_reset)

    output_dir = session_dir / "edited_outputs"
    output_dir.mkdir(parents=True, exist_ok=True)

    words = re.findall(r"[A-Za-z0-9]+", question.lower())[:6]
    base_name = "_".join(words) if words else "combined_output"
    base_name = re.sub(r"[^A-Za-z0-9_]", "", base_name).strip("_") or "combined_output"
    filename = f"{base_name}_combined.xlsx"

    output_path = output_dir / filename
    wb.save(str(output_path))

    return {
        "handled": True,
        "filename": output_path.name,
        "sheet_names": wb.sheetnames,
        "pivot_rows": len(result_df),
        "pivot_cols": len(result_df.columns),
        "original_rows": len(primary_df),
        "original_cols": len(primary_df.columns),
        "total_charts": 3,
        "preview": pivot_df_reset.head(10).to_dict(orient="records"),
        "columns": list(pivot_df_reset.columns.astype(str)),
    }


def run_excel_edit(
    session_dir: Path,
    question: str,
    base_url: str,
    model: str,
) -> Dict[str, Any]:
    cleaned_dir = preprocess_tabular_files(session_dir)
    if not cleaned_dir.exists():
        return {"handled": False}

    csv_files = [p for p in cleaned_dir.iterdir() if p.suffix.lower() == ".csv"]
    if not csv_files:
        return {"handled": False}

    preloaded_dfs: Dict[str, pd.DataFrame] = {}
    schema_parts: list[str] = []

    for csv_file in sorted(csv_files):
        try:
            df = pd.read_csv(csv_file)
        except Exception as exc:
            logger.error(f"Error preloading/parsing {csv_file.name}: {exc}")
            continue

        var_name = sanitize_df_name(csv_file.name)
        preloaded_dfs[var_name] = df
        cols = df.columns.tolist()
        sample_rows = df.head(2).to_string(index=False)

        schema_parts.append(
            f"- DataFrame variable: `{var_name}` (sourced from '{csv_file.name}')\n"
            f"  Columns: {cols}\n"
            f"  Sample Data:\n"
            f"  {sample_rows}"
        )

    if not preloaded_dfs:
        return {"handled": False}

    schema_str = "\n\n".join(schema_parts)

    prompt = f"""You are a professional Python data analyst. The user wants to modify or transform data.

The following pandas DataFrames are pre-loaded in the environment:
{schema_str}

Instructions:
- Write Python Pandas code to perform the requested transformation.
- The final result must be stored in a variable called exactly result_df (mandatory).
- If the operation is a pivot, result_df should be the pivot table (use pd.pivot_table or df.pivot_table), then call result_df.reset_index() so the pivot index becomes a regular column.
- If the operation adds or modifies a column, result_df should be the modified DataFrame.
- If the operation adds a blank row, append a new row with empty values for all columns.
- Do NOT use pd.read_csv or read any files from disk.
- Do NOT import anything (pd and np are already available).
- Do NOT call plt.show() or any plotting functions.
- Do NOT print anything; just assign result_df.
- Return only the Python code block starting with ```python and ending with ```.

User instruction: {question}
Python Code:"""

    from app.services.ollama_client import ollama_chat

    messages = [{"role": "user", "content": prompt}]
    success = False
    result_df: pd.DataFrame | None = None

    if _is_blank_row_request(question):
        try:
            target = _select_target_df(question, preloaded_dfs, csv_files)
            name_value = _extract_named_value(question)
            result_df = _append_blank_row(preloaded_dfs[target], name_value)
            success = True
        except Exception as exc:
            logger.error(f"Failed to append blank row deterministically: {exc}")

    for attempt in range(1, 4):
        if success:
            break
        code = ""
        try:
            text = ollama_chat(base_url, model, messages)

            code = text
            if "```python" in text:
                code = text.split("```python")[1].split("```")[0]

            _ensure_safe_code(code)

            exec_globals = {"pd": pd, "np": np, **preloaded_dfs}
            exec_locals: Dict[str, Any] = {}

            stdout_buffer = io.StringIO()
            with contextlib.redirect_stdout(stdout_buffer):
                _run_with_timeout(lambda: exec(code, exec_globals, exec_locals), 20)

            result_df = exec_locals.get("result_df")
            if result_df is None:
                result_df = exec_globals.get("result_df")

            if isinstance(result_df, pd.DataFrame):
                success = True
                break

            raise RuntimeError("result_df was not set to a pandas DataFrame.")
        except Exception:
            import traceback

            error_str = traceback.format_exc()
            logger.warning(
                f"Excel edit execution failed on attempt {attempt}. Error:\n{error_str}"
            )
            if code:
                messages.append({"role": "assistant", "content": f"```python\n{code}\n```"})
            error_msg = (
                "The code execution failed with the following traceback/error:\n"
                f"{error_str.strip()}\n\n"
                "Please correct the code. Ensure you use the correct preloaded DataFrame variables "
                "and write correct Pandas code. Return ONLY the executable python block."
            )
            messages.append({"role": "user", "content": error_msg})

    if not success or result_df is None:
        return {"handled": False}

    if isinstance(result_df.columns, pd.MultiIndex):
        result_df.columns = [
            "_".join(str(level) for level in col).strip("_")
            for col in result_df.columns
        ]

    if result_df.index.name is not None or isinstance(result_df.index, pd.MultiIndex):
        result_df = result_df.reset_index()

    output_dir = session_dir / "edited_outputs"
    output_dir.mkdir(parents=True, exist_ok=True)

    words = re.findall(r"[A-Za-z0-9]+", question.lower())[:6]
    base_name = "_".join(words) if words else "edited_output"
    base_name = re.sub(r"[^A-Za-z0-9_]", "", base_name).strip("_") or "edited_output"
    filename = f"{base_name}_edited.xlsx"

    output_path = output_dir / filename
    result_df.to_excel(output_path, index=False, engine="openpyxl")

    return {
        "handled": True,
        "filename": output_path.name,
        "row_count": len(result_df),
        "col_count": len(result_df.columns),
        "preview": result_df.head(10).to_dict(orient="records"),
        "columns": list(result_df.columns.astype(str)),
    }


def get_edited_file_path(session_dir: Path, filename: str) -> Path | None:
    safe_name = re.sub(r"[^A-Za-z0-9_\-.]", "", Path(filename).name)
    if not safe_name:
        return None
    file_path = session_dir / "edited_outputs" / safe_name
    return file_path if file_path.exists() else None
